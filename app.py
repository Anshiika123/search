from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
import requests
import re
import os
import io
import json
import urllib.parse
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
try:
    import wordninja
    _HAS_WORDNINJA = True
except ImportError:
    _HAS_WORDNINJA = False

load_dotenv()

app = Flask(__name__)
search_results = {}
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "")


def parse_keywords(topic: str) -> list:
    raw = re.split(r'[\s,&+/]+', topic.strip())
    stopwords = {'and','or','the','a','an','for','with','in','on','at','to'}
    return [w.lower() for w in raw if len(w) >= 2 and w.lower() not in stopwords]


def keyword_score(text: str, keywords: list) -> int:
    t = text.lower()
    return sum(1 for kw in keywords if kw in t)


def serpapi_search(query: str, api_key: str, num: int = 10) -> list:
    try:
        resp = requests.get("https://serpapi.com/search", params={
            "q": query, "api_key": api_key,
            "engine": "google", "num": num, "hl": "en",
        }, timeout=15)
        data = resp.json()
        if "error" in data:
            print(f"SerpAPI error: {data['error']}")
            return []
        return data.get("organic_results", [])
    except Exception as e:
        print(f"SerpAPI request failed: {e}")
        return []


def parse_profile_signals(snippet: str, title: str) -> dict:
    text = (title + " " + snippet).lower()
    raw  = title + " " + snippet

    open_to_work = any(p in text for p in [
        "open to work", "open to opportunities", "looking for opportunities",
        "seeking new role", "available for hire", "actively looking", "#opentowork"
    ])

    current_company = ""
    for pat in [
        r'\bat\s+([A-Z][A-Za-z0-9\s&.,]{2,30}?)(?:\s*[|·\-]|\s*$)',
        r'[@·]\s*([A-Z][A-Za-z0-9\s&.,]{2,25}?)(?:\s*[|·\-]|\s*$)',
    ]:
        m = re.search(pat, raw)
        if m:
            candidate = m.group(1).strip().rstrip('.,')
            if len(candidate) > 2 and candidate.lower() not in ('linkedin','the','and','for'):
                current_company = candidate
                break

    job_title = ""
    title_clean = re.sub(r'\s*\|\s*LinkedIn.*$', '', title, flags=re.I).strip()
    title_clean = re.sub(r'\s+on\s+LinkedIn.*$', '', title_clean, flags=re.I).strip()
    title_clean = re.sub(r'\s*-\s*LinkedIn.*$', '', title_clean, flags=re.I).strip()
    parts = re.split(r'\s*[-–|·]\s*', title_clean, maxsplit=1)
    if len(parts) >= 2:
        candidate = parts[1].strip()
        candidate = re.sub(r',\s*(MS|MBA|PhD|BSc|MSc|BE|BTech|MTech)\.?\s*\.{0,3}$', '', candidate, flags=re.I).strip()
        candidate = re.sub(r'\s*\.{2,}$', '', candidate).strip()
        if len(candidate) > 3:
            job_title = candidate

    seniority_score = 0
    seniority_kws = {
        5: ['cto','ceo','coo','chief','vp ','vice president','founder','co-founder'],
        4: ['director','head of','principal','distinguished','fellow'],
        3: ['senior','sr.','lead','staff','architect','manager'],
        2: ['associate','consultant','specialist','analyst'],
        1: ['junior','jr.','intern','trainee','fresher','student'],
    }
    for score, kws in seniority_kws.items():
        if any(kw in text for kw in kws):
            seniority_score = score
            break

    exp_years = 0
    m = re.search(r'(\d{1,2})\+?\s*(?:years?|yrs?)\s*(?:of\s*)?(?:experience|exp)', text)
    if m:
        exp_years = int(m.group(1))
        seniority_score = max(seniority_score, min(5, exp_years // 3))

    is_active = bool(re.search(
        r'\b(\d+[hm]\s*ago|today|yesterday|just now|\d+\s*days?\s*ago|this week)\b', text
    ))

    portfolio_checks = {
        "GitHub":      ["github", "github.com"],
        "Open Source": ["open source", "opensource", "contributor"],
        "Built":       ["built ", "i built", "developed ", "created "],
        "Launched":    ["launched", "shipped", "released"],
        "Patent":      ["patent", "patented", "inventor"],
        "Speaker":     ["speaker", "keynote", "ted talk"],
        "Published":   ["published", "author of", "research paper"],
    }
    portfolio_signals = [label for label, kws in portfolio_checks.items() if any(kw in text for kw in kws)]

    # Extract current location from snippet
    location = ""
    # Pattern 1: "City, State, Country" or "City, State" separated by · or commas
    loc_patterns = [
        r'·\s*([A-Z][A-Za-z\s]+,\s*[A-Z][A-Za-z\s]+(?:,\s*[A-Z][A-Za-z\s]+)?)\s*·',
        r'Location[:\s]+([A-Z][A-Za-z\s]+(?:,\s*[A-Z][A-Za-z\s]+){0,2})',
        r'\b([A-Z][A-Za-z]+(?:\s[A-Z][A-Za-z]+)?,\s*(?:Karnataka|Maharashtra|Tamil Nadu|Telangana|Delhi|Gujarat|Rajasthan|UP|West Bengal|Punjab|Haryana|Kerala|Andhra Pradesh|India))\b',
        r'\b([A-Z][A-Za-z\s]+,\s*(?:USA|UK|UAE|Canada|Australia|Singapore|Germany|Netherlands|France))\b',
    ]
    for pat in loc_patterns:
        lm = re.search(pat, raw)
        if lm:
            candidate = lm.group(1).strip().strip('·').strip()
            if 3 < len(candidate) < 60:
                location = candidate
                break

    return {
        "current_company": current_company,
        "job_title": job_title,
        "open_to_work": open_to_work,
        "seniority_score": seniority_score,
        "exp_years": exp_years,
        "is_active": is_active,
        "portfolio_signals": portfolio_signals,
        "portfolio_score": len(portfolio_signals),
        "location": location,
    }


def clean_author_name(raw: str) -> str:
    if not raw:
        return ""
    raw = re.sub(r'[,|·–]\s*(MS|MBA|PhD|BSc|MSc|BE|BTech|MTech|CPA|CFA)\.?.*$', '', raw, flags=re.I).strip()
    raw = re.sub(r'[|·–].*$', '', raw).strip()
    if ' ' not in raw and '-' in raw:
        raw = raw.replace('-', ' ')
    if ' ' not in raw:
        # Try camelCase split first
        split = re.sub(r'([a-z])([A-Z])', r'\1 \2', raw)
        if ' ' in split:
            raw = split
        elif _HAS_WORDNINJA:
            # Slug like "aaronchall" → ["aaron", "chall"]
            parts = wordninja.split(raw.lower())
            if 2 <= len(parts) <= 4:
                raw = ' '.join(parts)
    words = raw.split()
    stop = {'a','an','the','of','at','in','on','and','or','for','to'}
    name = ' '.join(w.capitalize() if w.lower() not in stop or i == 0 else w
                    for i, w in enumerate(words))
    return name.strip()


_JOB_WORDS = {
    'analyst', 'developer', 'engineer', 'manager', 'director', 'designer',
    'consultant', 'specialist', 'architect', 'researcher', 'scientist',
    'lead', 'senior', 'junior', 'intern', 'head', 'vp', 'cto', 'ceo',
    'founder', 'officer', 'associate', 'assistant', 'student', 'aspiring',
    'graduate', 'faculty', 'professor', 'teacher', 'trainer', 'recruiter',
    'data', 'software', 'product', 'project', 'business', 'marketing',
    'sales', 'operations', 'strategy', 'finance', 'accounting', 'talent',
    'cloud', 'devops', 'security', 'fullstack', 'frontend', 'backend',
    'full', 'stack', 'freelance', 'freelancer', 'remote', 'contract',
}
# Tech keywords that should never be the first word of a person's name
_TECH_FIRST_WORDS = {
    'python', 'java', 'javascript', 'typescript', 'ruby', 'golang', 'rust',
    'react', 'angular', 'vue', 'node', 'django', 'flask', 'spring',
    'android', 'ios', 'mobile', 'web', 'machine', 'deep', 'artificial',
    'blockchain', 'crypto', 'aws', 'azure', 'gcp', 'linux', 'docker',
    'kubernetes', 'sql', 'nosql', 'mongodb', 'postgres', 'mysql',
}

def _looks_like_name(s: str) -> bool:
    words = s.strip().split()
    if not (2 <= len(words) <= 4):       # real names have at least 2 words
        return False
    if not all(w and w[0].isupper() for w in words):
        return False
    # Reject all-caps abbreviations (SQL AI ML…)
    if all(w.isupper() and len(w) <= 4 for w in words):
        return False
    # Reject if first word is a tech keyword
    if words[0].lower() in _TECH_FIRST_WORDS:
        return False
    # Reject if any word is a job/role word
    if {w.lower() for w in words} & _JOB_WORDS:
        return False
    return True


def extract_author(url: str, title: str) -> tuple:
    profile_url = ""
    raw_name = ""
    clean = re.sub(r'<[^>]+>', '', title).strip()

    # Strip LinkedIn suffix variants to get the name portion
    before_linkedin = ""
    for pat in [" on LinkedIn", " | LinkedIn", " - LinkedIn"]:
        if pat in clean:
            before_linkedin = clean.split(pat)[0].strip()
            break
    if not before_linkedin and " - " in clean and "LinkedIn" in clean:
        before_linkedin = clean.split(" - ")[0].strip()

    # Split into segments and find the one that looks like a person's name
    if before_linkedin:
        segments = re.split(r'\s*[-–|·]\s*', before_linkedin)
        for seg in segments:
            seg = seg.strip()
            if _looks_like_name(seg):
                raw_name = seg
                break
        # If no segment looks like a name, leave raw_name empty — slug fallback below

    slug_m = re.search(r'linkedin\.com/in/([a-zA-Z0-9][a-zA-Z0-9\-]+)', url)
    if slug_m:
        clean_slug = re.sub(r'-[a-z0-9]*\d{4,}[a-z0-9]*$', '', slug_m.group(1), flags=re.I).strip('-')
        if clean_slug:
            profile_url = f"https://www.linkedin.com/in/{clean_slug}"
            if not raw_name:
                raw_name = clean_slug

    slug_m2 = re.search(r'linkedin\.com/posts/([a-zA-Z0-9][a-zA-Z0-9\-]+)', url)
    if slug_m2:
        clean_slug = re.sub(r'-[a-z0-9]*\d{4,}[a-z0-9]*$', '', slug_m2.group(1), flags=re.I).strip('-')
        if clean_slug and not profile_url:
            profile_url = f"https://www.linkedin.com/in/{clean_slug}"
        if not raw_name:
            raw_name = clean_slug

    author_name = clean_author_name(raw_name)
    # Reject names that are clearly job titles / tech keywords, not real people
    if author_name:
        words = author_name.lower().split()
        if (words[0] in _TECH_FIRST_WORDS or
                bool({w for w in words} & _JOB_WORDS)):
            return "", profile_url
    return author_name, profile_url


# Common alternate spellings / nearby areas for popular regions
_REGION_ALIASES = {
    "bangalore": ["Bangalore", "Bengaluru", "Bengaluru Karnataka"],
    "bengaluru": ["Bengaluru", "Bangalore", "Bengaluru Karnataka"],
    "bombay":    ["Mumbai", "Bombay"],
    "mumbai":    ["Mumbai", "Bombay"],
    "delhi":     ["Delhi", "New Delhi", "NCR"],
    "ncr":       ["NCR", "Delhi", "Noida", "Gurugram", "Gurgaon"],
    "madras":    ["Chennai", "Madras"],
    "chennai":   ["Chennai", "Madras"],
    "calcutta":  ["Kolkata", "Calcutta"],
    "kolkata":   ["Kolkata", "Calcutta"],
    "hyderabad": ["Hyderabad", "Secunderabad", "Cyberabad"],
    "pune":      ["Pune", "Pimpri"],
    "noida":     ["Noida", "Greater Noida", "NCR"],
    "gurugram":  ["Gurugram", "Gurgaon", "NCR"],
    "gurgaon":   ["Gurgaon", "Gurugram", "NCR"],
}

def _region_variants(region_q: str) -> list:
    """Return quoted region variants to use in queries."""
    key = region_q.lower().split()[0]
    aliases = _REGION_ALIASES.get(key, [region_q])
    # Always include the original too
    if region_q not in aliases:
        aliases = [region_q] + aliases
    return [f'"{a}"' for a in aliases[:3]]  # quoted for exact match


def build_profile_queries(kw: str, keywords: list, region_q: str, status_filter: str) -> list:
    status_terms = {
        "open_to_work":  ["open to work", "seeking opportunities"],
        "experienced":   ["senior", "lead", "5 years", "10 years", "architect"],
    }.get(status_filter, [])

    queries = []
    if region_q:
        for region_variant in _region_variants(region_q):
            queries.append(f'site:linkedin.com/in/ {kw} {region_variant}')
        if status_terms:
            queries.append(f'site:linkedin.com/in/ {kw} {status_terms[0]} {_region_variants(region_q)[0]}')
        for kw_single in keywords[:2]:
            queries.append(f'site:linkedin.com/in/ {kw_single} {_region_variants(region_q)[0]}')

    queries.append(f'site:linkedin.com/in/ {kw}')
    if status_terms:
        for term in status_terms[:2]:
            queries.append(f'site:linkedin.com/in/ {kw} {term}')
    for kw_single in keywords[:3]:
        queries.append(f'site:linkedin.com/in/ {kw_single}')

    seen_q, unique = set(), []
    for q in queries:
        if q not in seen_q:
            seen_q.add(q)
            unique.append(q)
    return unique


def search_profiles(keywords, region, api_key, fetch_size, status_filter="all"):
    kw = " ".join(keywords)
    region_parts = [p for p in re.split(r'[\s,]+', region) if len(p) > 2] if region else []
    region_q = " ".join(region_parts[:2]) if region_parts else ""

    queries = build_profile_queries(kw, keywords, region_q, status_filter)
    candidates = []
    seen = set()

    for query in queries:
        items = serpapi_search(query, api_key, num=10)
        for item in items:
            url     = item.get("link", "")
            title   = item.get("title", "")
            snippet = item.get("snippet", "")

            if "linkedin.com/in/" not in url and "linkedin.com/in/" not in url.replace("in.linkedin", "linkedin"):
                continue
            url = url.replace("https://in.linkedin.com", "https://www.linkedin.com")
            url = url.split("?")[0].rstrip("/")
            if url in seen:
                continue
            seen.add(url)

            author, profile_url = extract_author(url, title)
            if not author:
                continue  # skip anonymous profiles

            full = (title + " " + snippet).lower()
            signals = parse_profile_signals(snippet, title)
            region_terms = set(region_parts)
            for rp in region_parts:
                for alias in _REGION_ALIASES.get(rp.lower(), []):
                    region_terms.add(alias.lower())
            region_match = any(t.lower() in full for t in region_terms) if region_terms else False

            score = keyword_score(full, keywords)
            if region_match:              score += 3
            score += signals["seniority_score"]
            score += signals["portfolio_score"]
            if signals["exp_years"] >= 5: score += 2
            elif signals["exp_years"] >= 2: score += 1
            if signals["is_active"]:      score += 1
            if signals["open_to_work"]:   score += 1

            post_search = f"https://www.linkedin.com/search/results/content/?keywords={urllib.parse.quote(author)}"
            candidates.append((score, {
                "author_name": author,
                "profile_url": profile_url or url,
                "post_url": post_search,
                "post_url_label": "🔍 Search Their Posts",
                "context": snippet[:400],
                "matched_keywords": keyword_score(full, keywords),
                "total_keywords": len(keywords),
                "region_match": region_match,
                "result_type": "profile",
                "current_company": signals["current_company"],
                "job_title": signals["job_title"],
                "open_to_work": signals["open_to_work"],
                "seniority_score": signals["seniority_score"],
                "exp_years": signals["exp_years"],
                "is_active": signals["is_active"],
                "portfolio_signals": signals["portfolio_signals"],
                "location": signals["location"],
            }))

    return candidates


def search_posts(keywords, region, api_key):
    kw = " ".join(keywords)
    region_parts = [p for p in re.split(r'[\s,]+', region) if len(p) > 2] if region else []
    region_q = " ".join(region_parts[:2]) if region_parts else ""

    queries = []
    if region_q:
        queries.append(f'site:linkedin.com/posts/ {kw} {region_q}')
    queries.append(f'site:linkedin.com/posts/ {kw}')
    queries.append(f'site:linkedin.com/feed/update/ {kw}')
    for kw_single in keywords[:2]:
        queries.append(f'site:linkedin.com/posts/ {kw_single}')

    candidates = []
    seen = set()

    for query in queries:
        items = serpapi_search(query, api_key, num=10)
        for item in items:
            url     = item.get("link", "")
            title   = item.get("title", "")
            snippet = item.get("snippet", "")

            if "linkedin.com/posts/" not in url and "linkedin.com/feed/update/" not in url:
                continue
            url = url.replace("https://in.linkedin.com", "https://www.linkedin.com")
            url = url.split("?")[0].rstrip("/")
            if url in seen:
                continue
            seen.add(url)

            author, profile_url = extract_author(url, title)
            if not author:
                continue

            full = (title + " " + snippet).lower()
            signals = parse_profile_signals(snippet, title)
            region_terms = set(region_parts)
            for rp in region_parts:
                for alias in _REGION_ALIASES.get(rp.lower(), []):
                    region_terms.add(alias.lower())
            region_match = any(t.lower() in full for t in region_terms) if region_terms else False

            score = keyword_score(full, keywords)
            if region_match:         score += 2
            if signals["is_active"]: score += 1

            candidates.append((score, {
                "author_name": author,
                "profile_url": profile_url,
                "post_url": url,
                "post_url_label": "📄 Open Post",
                "context": snippet[:400],
                "matched_keywords": keyword_score(full, keywords),
                "total_keywords": len(keywords),
                "region_match": region_match,
                "result_type": "post",
                "current_company": signals["current_company"],
                "job_title": signals["job_title"],
                "open_to_work": signals["open_to_work"],
                "seniority_score": signals["seniority_score"],
                "exp_years": signals["exp_years"],
                "is_active": signals["is_active"],
                "portfolio_signals": signals["portfolio_signals"],
                "location": signals["location"],
            }))

    return candidates


def search_linkedin(topic, max_results=5, region="", verified_only=False,
                    search_mode="both", api_key="", status_filter="all"):
    api_key = api_key or SERPAPI_KEY
    if not api_key:
        return {"error": "SerpAPI key missing. Please enter it below."}

    keywords = parse_keywords(topic)
    if not keywords:
        return {"error": "Please enter a topic."}

    fetch_size = max(50, max_results * 5)
    region_parts = [p for p in re.split(r'[\s,]+', region) if len(p) > 2] if region else []

    candidates = []
    if search_mode in ("profiles", "both"):
        candidates += search_profiles(keywords, region, api_key, fetch_size, status_filter)
    if search_mode in ("posts", "both"):
        candidates += search_posts(keywords, region, api_key)

    if not candidates:
        msg = f"No results found for '{topic}'"
        if region:
            msg += f" in '{region}'. Try broader keywords or leave Region empty."
        else:
            msg += ". Try different keywords."
        return {"error": msg}

    candidates.sort(key=lambda x: x[0], reverse=True)

    # Deduplicate — type-specific key to avoid collision
    seen = set()
    all_posts = []
    for _, p in candidates:
        key = p.get("profile_url") if p.get("result_type") == "profile" else p.get("post_url", "")
        if not key or key in seen:
            continue
        seen.add(key)
        all_posts.append(p)

    # Region filter — strict: only show in-region results when a region is set
    region_filter_applied = False
    if region_parts:
        in_region = [p for p in all_posts if p.get("region_match")]
        if in_region:
            region_filter_applied = True
            all_posts = in_region  # never pad with non-region results

    # Status filter
    if status_filter == "open_to_work":
        filtered = [p for p in all_posts if p.get("open_to_work")]
    elif status_filter == "working":
        filtered = [p for p in all_posts if p.get("current_company") and not p.get("open_to_work")]
    elif status_filter == "experienced":
        filtered = [p for p in all_posts if p.get("exp_years", 0) >= 1 or p.get("seniority_score", 0) >= 1]
    else:
        filtered = all_posts

    # Supplement if filter too strict
    if len(filtered) < max_results and status_filter != "all":
        extras = [p for p in all_posts if p not in filtered]
        filtered = filtered + extras

    posts = filtered[:max_results]

    return {
        "posts": posts,
        "topic": topic,
        "count": len(posts),
        "total_pool": len(all_posts),
        "filtered_total": len(filtered),
        "keywords": keywords,
        "region": region,
        "region_filter_applied": region_filter_applied,
        "verified_only": verified_only,
        "search_mode": search_mode,
        "status_filter": status_filter,
    }


@app.route("/debug")
def debug():
    api_key = SERPAPI_KEY
    query = request.args.get("q", "site:linkedin.com/in/ python Bangalore")
    try:
        resp = requests.get("https://serpapi.com/search", params={
            "q": query, "api_key": api_key, "engine": "google", "num": 5
        }, timeout=15)
        data = resp.json()
    except Exception as e:
        return f"<pre>Error: {e}</pre>"
    if "error" in data:
        return f"<h3>{query}</h3><p style='color:red'>Error: {data['error']}</p>"
    items = data.get("organic_results", [])
    out = f"<h3>{query}</h3><p>{len(items)} results</p>"
    for i, item in enumerate(items, 1):
        out += f"<hr><b>#{i}</b> {item.get('title','')}<br>"
        out += f"<a href='{item.get('link','')}' target='_blank'>{item.get('link','')}</a><br>"
        out += f"{item.get('snippet','')}"
    return out


@app.route("/")
def index():
    return render_template("index.html", has_creds=bool(SERPAPI_KEY))


@app.route("/search", methods=["POST"])
def search():
    data = request.json
    topic = data.get("topic", "").strip()
    if not topic:
        return jsonify({"error": "Please enter a topic."})
    try:
        result = search_linkedin(
            topic,
            max_results=int(data.get("max_results", 5)),
            region=data.get("region", "").strip(),
            verified_only=bool(data.get("verified_only", False)),
            search_mode=data.get("search_mode", "both"),
            api_key=data.get("api_key", "").strip(),
            status_filter=data.get("status_filter", "all"),
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": f"Server error: {str(e)}"})
    if "posts" in result:
        search_results["last"] = result
    return jsonify(result)


@app.route("/save_creds", methods=["POST"])
def save_creds():
    data = request.json
    key  = data.get("api_key", "").strip()
    if not key:
        return jsonify({"error": "Key required."})
    env_path = Path(__file__).parent / ".env"
    env_path.write_text(f"SERPAPI_KEY={key}\n")
    global SERPAPI_KEY
    SERPAPI_KEY = key
    return jsonify({"ok": True})


@app.route("/export")
def export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl not installed", 500

    data = search_results.get("last")
    if not data:
        return "No results to export", 400

    posts  = data["posts"]
    topic  = data["topic"]
    region = data.get("region", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Results"
    hf    = PatternFill("solid", fgColor="0A66C2")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    thin  = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_val = f'LinkedIn: "{topic}"' + (f' | {region}' if region else '') + f' -- {len(posts)} results'
    ws.merge_cells("A1:G1")
    ws["A1"].value = title_val
    ws["A1"].font  = Font(bold=True, size=13, color="0A66C2")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws["A2"].font  = Font(italic=True, color="888888", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    cols   = ["#","Type","Author","Profile Link","Post / Search Link","Snippet","Region Match"]
    widths = [5, 10, 22, 40, 42, 55, 14]
    for c,(h,w) in enumerate(zip(cols,widths),1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill=hf; cell.font=hfont; cell.border=border
        cell.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[cell.column_letter].width=w
    ws.row_dimensions[3].height=20

    for i,post in enumerate(posts,1):
        row=i+3
        alt=PatternFill("solid",fgColor="EBF3FB" if i%2==0 else "FFFFFF")
        vals=[i,
              "Profile" if post.get("result_type")=="profile" else "Post",
              post.get("author_name",""),
              post.get("profile_url",""),
              post.get("post_url",""),
              post.get("context",""),
              "Y" if post.get("region_match") else "-"]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=row,column=c,value=v)
            cell.fill=alt; cell.border=border
            cell.alignment=Alignment(vertical="top",wrap_text=True)
            if c in(4,5) and v: cell.font=Font(color="0A66C2",underline="single")
        ws.row_dimensions[row].height=55

    ws.freeze_panes="A4"
    safe=re.sub(r'[^\w\s-]','',topic).strip().replace(' ','_')
    fname=f"LinkedIn_{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    buf=io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




@app.route("/search_stream")
def search_stream():
    topic         = request.args.get("topic", "").strip()
    region        = request.args.get("region", "").strip()
    max_results   = int(request.args.get("max_results", 5))
    search_mode   = request.args.get("search_mode", "both")
    status_filter = request.args.get("status_filter", "all")
    api_key       = request.args.get("api_key", "").strip() or SERPAPI_KEY

    def generate():
        if not topic or not api_key:
            yield "data: " + json.dumps({"type": "error", "msg": "Missing topic or API key"}) + "\n\n"
            return

        keywords     = parse_keywords(topic)
        region_parts = [p for p in re.split(r"[\s,]+", region) if len(p) > 2] if region else []
        region_q     = " ".join(region_parts[:2]) if region_parts else ""
        seen         = set()
        candidates   = []
        kw           = " ".join(keywords)

        def process_item(item, result_type):
            url     = item.get("link", "")
            title   = item.get("title", "")
            snippet = item.get("snippet", "")
            if result_type == "profile":
                if "linkedin.com/in/" not in url:
                    return None
            else:
                if "linkedin.com/posts/" not in url and "linkedin.com/feed/update/" not in url:
                    return None
            url = url.replace("https://in.linkedin.com", "https://www.linkedin.com").split("?")[0].rstrip("/")
            if url in seen:
                return None
            seen.add(url)
            author, profile_url = extract_author(url, title)
            if not author:
                return None
            full         = (title + " " + snippet).lower()
            signals      = parse_profile_signals(snippet, title)
            region_terms = set(region_parts)
            for rp in region_parts:
                for alias in _REGION_ALIASES.get(rp.lower(), []):
                    region_terms.add(alias.lower())
            region_match = any(t.lower() in full for t in region_terms) if region_terms else False
            score = keyword_score(full, keywords)
            if region_match:              score += 3
            score += signals["seniority_score"]
            score += signals["portfolio_score"]
            if signals["exp_years"] >= 5: score += 2
            elif signals["exp_years"] >= 2: score += 1
            if signals["is_active"]:      score += 1
            if signals["open_to_work"]:   score += 1
            post_search = "https://www.linkedin.com/search/results/content/?keywords=" + urllib.parse.quote(author)
            return {
                "author_name":       author,
                "profile_url":       profile_url or url,
                "post_url":          url if result_type == "post" else post_search,
                "context":           snippet[:400],
                "matched_keywords":  keyword_score(full, keywords),
                "total_keywords":    len(keywords),
                "region_match":      region_match,
                "result_type":       result_type,
                "current_company":   signals["current_company"],
                "job_title":         signals["job_title"],
                "open_to_work":      signals["open_to_work"],
                "seniority_score":   signals["seniority_score"],
                "exp_years":         signals["exp_years"],
                "is_active":         signals["is_active"],
                "portfolio_signals": signals["portfolio_signals"],
                "location":          signals["location"],
                "_score":            score,
            }

        queries = []
        if search_mode in ("profiles", "both"):
            if region_q:
                for rv in _region_variants(region_q):
                    queries.append(("profile", "site:linkedin.com/in/ " + kw + " " + rv))
                for k in keywords[:2]:
                    queries.append(("profile", "site:linkedin.com/in/ " + k + " " + _region_variants(region_q)[0]))
            queries.append(("profile", "site:linkedin.com/in/ " + kw))
        if search_mode in ("posts", "both"):
            if region_q:
                queries.append(("post", "site:linkedin.com/posts/ " + kw + " " + _region_variants(region_q)[0]))
            queries.append(("post", "site:linkedin.com/posts/ " + kw))

        total_q = len(queries)
        for qi, (rtype, query) in enumerate(queries):
            yield "data: " + json.dumps({"type": "progress", "step": qi+1, "total": total_q, "query": query}) + "\n\n"
            items = serpapi_search(query, api_key, num=10)
            for item in items:
                p = process_item(item, rtype)
                if p:
                    candidates.append(p)
                    yield "data: " + json.dumps({"type": "result", "profile": p}) + "\n\n"

        candidates.sort(key=lambda x: x.get("_score", 0), reverse=True)

        # Strict region filter — only show in-region results when region is set
        region_filter_applied = False
        pool = candidates
        if region_parts:
            in_region = [p for p in candidates if p.get("region_match")]
            if in_region:
                region_filter_applied = True
                pool = in_region

        final = pool[:max_results]
        for p in final:
            p.pop("_score", None)
        yield "data: " + json.dumps({
            "type": "done", "ranked": final, "total_pool": len(pool),
            "topic": topic, "region": region, "keywords": keywords,
            "status_filter": status_filter, "region_filter_applied": region_filter_applied,
        }) + "\n\n"

    return Response(stream_with_context(generate()), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
